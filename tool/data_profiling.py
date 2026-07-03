import pandas as pd
import numpy as np
import re
import argparse
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class DataProfiler:
    """
    Data profiling theo 3 bước chuẩn (Structure / Content / Relationship Discovery):
    - Content Discovery: NULL = giá trị thiếu THẬT SỰ (NaN hoặc chuỗi rỗng/khoảng trắng)
      -> áp dụng NHẤT QUÁN cho mọi thống kê (Cardinality, Mode, Pattern, Outlier...),
      không chỉ riêng bảng NULL/Actual.
    - Structure Discovery: tự suy luận kiểu dữ liệu thực sự của cột (Integer/Decimal/Date/String)
      kể cả khi cột có lẫn vài giá trị rác (vd Quantity = "two", "1.5"), thay vì chỉ dựa vào
      dtype gốc của pandas (dễ bị "gãy" thành object chỉ vì 1 giá trị lỗi).
    - Validity: % giá trị hợp lệ theo regex do người dùng cung cấp, HOẶC (nếu không có regex)
      theo tỉ lệ parse thành công khi ép kiểu số/ngày - đúng tinh thần bảng Validity trong tài liệu.
    """

    def __init__(self, df):
        self.df = df
        self.total_rows = len(df)
        self._missing_masks = {}
        self._actual_cache = {}
        for col in self.df.columns:
            s = self.df[col]
            if pd.api.types.is_numeric_dtype(s) or pd.api.types.is_datetime64_any_dtype(s):
                mask = s.isnull()
            else:
                mask = s.isnull() | (s.astype(str).str.strip() == '')
            self._missing_masks[col] = mask
            self._actual_cache[col] = s[~mask]

    def _numeric_series(self, col):
        """Ép kiểu số cho phần dữ liệu Actual (không NULL). Trả về (series đã ép, tỉ lệ hợp lệ).
        Dùng sample trước để tránh chạy regex/parse tốn thời gian trên các cột text dài
        (vd 'comments' hàng triệu dòng) khi rõ ràng không phải cột số."""
        actual = self._actual_cache[col]
        n = len(actual)
        if n == 0:
            return pd.Series(dtype=float), None

        if pd.api.types.is_numeric_dtype(actual):
            return actual, 1.0

        sample = actual if n <= 2000 else actual.sample(2000, random_state=0)
        sample_direct = pd.to_numeric(sample, errors='coerce')
        sample_validity = sample_direct.notna().mean()

        if sample_validity < 0.3:
            # Thử làm sạch sample xem có phải số bị dính ký tự ($, %) không
            sample_cleaned_str = sample.astype(str).str.replace(r'[^0-9eE.\-+]', '', regex=True).replace('', np.nan)
            sample_cleaned = pd.to_numeric(sample_cleaned_str, errors='coerce')
            if sample_cleaned.notna().mean() < 0.3:
                # Không có dấu hiệu là cột số kể cả sau khi làm sạch -> bỏ qua
                return pd.Series(dtype=float), sample_validity

        direct = pd.to_numeric(actual, errors='coerce')
        validity = direct.notna().sum() / n
        if validity >= 0.98:
            return direct, validity

        # còn sót giá trị dạng "$300", "1,200 USD" -> làm sạch phần chưa parse được rồi ép lại
        unresolved_mask = direct.isna()
        if unresolved_mask.any():
            to_clean = actual[unresolved_mask].astype(str).str.replace(r'[^0-9eE.\-+]', '', regex=True)
            to_clean = to_clean.replace('', np.nan)
            cleaned = pd.to_numeric(to_clean, errors='coerce')
            direct = direct.combine_first(cleaned)
            validity = direct.notna().sum() / n
        return direct, validity

    def _date_series(self, col):
        actual = self._actual_cache[col]
        n = len(actual)
        if n == 0:
            return pd.Series(dtype='datetime64[ns]'), None
        if pd.api.types.is_numeric_dtype(actual):
            return pd.Series(dtype='datetime64[ns]'), 0.0

        sample = actual if n <= 2000 else actual.sample(2000, random_state=0)
        sample_coerced = pd.to_datetime(sample, errors='coerce', format='mixed')
        sample_validity = sample_coerced.notna().mean()

        if sample_validity < 0.5:
            # Không có dấu hiệu là cột ngày tháng (vd văn bản tự do 'comments')
            # -> không parse toàn bộ hàng triệu dòng cho tốn thời gian.
            return pd.Series(dtype='datetime64[ns]'), sample_validity

        coerced = pd.to_datetime(actual, errors='coerce', format='mixed')
        validity = coerced.notna().sum() / n
        return coerced, validity

    def _friendly_type(self, col, is_numeric_like, numeric_coerced, is_date_like):
        s = self.df[col]
        if pd.api.types.is_integer_dtype(s):
            return 'Integer'
        if pd.api.types.is_float_dtype(s):
            return 'Decimal'
        if pd.api.types.is_datetime64_any_dtype(s):
            return 'Date'
        if is_numeric_like:
            vals = numeric_coerced.dropna()
            if len(vals) and (vals % 1 == 0).all():
                return 'Integer'
            return 'Decimal'
        if is_date_like:
            return 'Date'
        return 'String'

    def get_basic_stats(self):
        stats = []
        for col in self.df.columns:
            mask = self._missing_masks[col]
            actual_data = self._actual_cache[col]
            null_count = int(mask.sum())
            actual = self.total_rows - null_count
            completeness = actual / self.total_rows if self.total_rows > 0 else 0
            cardinality = actual_data.nunique()
            uniqueness = cardinality / self.total_rows if self.total_rows > 0 else 0
            distinctness = cardinality / actual if actual > 0 else 0

            stats.append({
                'Field Name': col,
                'NULL': null_count,
                'Actual': actual,
                'Completeness': completeness,
                'Cardinality': cardinality,
                'Uniqueness': uniqueness,
                'Distinctness': distinctness
            })
        return pd.DataFrame(stats)

    def get_numeric_stats(self, regex_dict=None):
        if regex_dict is None:
            regex_dict = {}
        stats = []
        for col in self.df.columns:
            actual_data = self._actual_cache[col]
            regex_pattern = regex_dict.get(col, None)

            numeric_coerced, numeric_validity = self._numeric_series(col)
            is_numeric_like = numeric_validity is not None and numeric_validity >= 0.5

            date_coerced, date_validity = (pd.Series(dtype='datetime64[ns]'), None)
            if not is_numeric_like:
                date_coerced, date_validity = self._date_series(col)
            is_date_like = date_validity is not None and date_validity >= 0.5

            if regex_pattern and len(actual_data) > 0:
                validity = actual_data.astype(str).str.match(regex_pattern).mean()
            elif is_numeric_like:
                validity = numeric_validity
            elif is_date_like:
                validity = date_validity
            else:
                validity = None

            min_val = max_val = avg_val = median_val = None
            if is_numeric_like:
                vals = numeric_coerced.dropna()
                if len(vals):
                    min_val, max_val = vals.min(), vals.max()
                    avg_val, median_val = vals.mean(), vals.median()
            elif is_date_like:
                vals = date_coerced.dropna()
                if len(vals):
                    min_val, max_val = vals.min().date(), vals.max().date()

            if is_numeric_like:
                mode_val = numeric_coerced.mode().iloc[0] if len(numeric_coerced.dropna()) else None
            elif is_date_like:
                mode_val = date_coerced.mode().iloc[0].date() if len(date_coerced.dropna()) else None
            else:
                mode_val = actual_data.mode().iloc[0] if len(actual_data) else None

            stats.append({
                'Field Name': col,
                'Regex': regex_pattern if regex_pattern else "",
                'Validity': validity,
                'Min': min_val,
                'Max': max_val,
                'Mode': mode_val,
                'AVG': avg_val,
                'Median': median_val
            })
        return pd.DataFrame(stats)

    def get_data_abstraction(self):
        stats = []
        for col in self.df.columns:
            actual_data = self._actual_cache[col]
            n = len(actual_data)
            
            numeric_coerced, numeric_validity = self._numeric_series(col)
            is_numeric_like = numeric_validity is not None and numeric_validity >= 0.5
            date_coerced, date_validity = (pd.Series(dtype='datetime64[ns]'), None)
            if not is_numeric_like:
                date_coerced, date_validity = self._date_series(col)
            is_date_like = date_validity is not None and date_validity >= 0.5

            # 1. Type (C, O, Q)
            if is_numeric_like or is_date_like:
                type_val = 'Q'
            else:
                type_val = 'C'
            
            # 2. Key/Value
            is_unique = (actual_data.nunique() / n > 0.95) if n > 0 else False
            is_id_name = 'id' in col.lower()
            if is_unique and is_id_name:
                key_value = 'K'
            else:
                key_value = 'V'
                
            # 3. Direction
            direction = 'Sequential' if type_val == 'Q' else ''
            
            # 4. Hierarchical
            hierarchical = 'D - M - Y' if is_date_like else ''
            
            # 5. Continuous/Discrete
            if is_numeric_like:
                # Check if it's float or int
                is_float = (numeric_coerced.dropna() % 1 != 0).any()
                cont_disc = 'Continuous' if is_float else 'Discrete'
            elif is_date_like:
                cont_disc = 'Continuous'
            else:
                cont_disc = 'Discrete'
                
            stats.append({
                'Attribute Name': col,
                'Type (C,O,Q)': type_val,
                'Key/Value': key_value,
                'Direction': direction,
                'Hierarchical': hierarchical,
                'Continuous/Discrete': cont_disc,
                'Semantic': col
            })
        return pd.DataFrame(stats)

    def get_pattern_and_outliers(self):
        stats = []
        for col in self.df.columns:
            actual_data = self._actual_cache[col]
            numeric_coerced, numeric_validity = self._numeric_series(col)
            is_numeric_like = numeric_validity is not None and numeric_validity >= 0.5
            date_coerced, date_validity = (pd.Series(dtype='datetime64[ns]'), None)
            if not is_numeric_like:
                date_coerced, date_validity = self._date_series(col)
            is_date_like = date_validity is not None and date_validity >= 0.5

            friendly_type = self._friendly_type(col, is_numeric_like, numeric_coerced, is_date_like)

            patterns = []
            if len(actual_data) > 0:
                def get_pattern(val):
                    v = str(val)
                    v = re.sub(r'[A-Za-z]', 's', v)
                    v = re.sub(r'[0-9]', 'd', v)
                    return v
                pattern_counts = actual_data.apply(get_pattern).value_counts(normalize=True).head(5)
                patterns = [f"{idx}: {val:.1%}" for idx, val in pattern_counts.items()]

            outliers = []
            if is_numeric_like:
                vals = numeric_coerced.dropna()
                if len(vals) > 0:
                    Q1 = vals.quantile(0.25)
                    Q3 = vals.quantile(0.75)
                    IQR = Q3 - Q1
                    outlier_values = vals[(vals < Q1 - 1.5 * IQR) | (vals > Q3 + 1.5 * IQR)]
                    outliers = outlier_values.unique()[:10].tolist()

            stats.append({
                'Field Name': col,
                'Data type': friendly_type,
                'Top 5 patterns': "\n".join(patterns),
                'Outliers (IQR, top 10)': ", ".join(map(str, outliers)) if outliers else ""
            })
        return pd.DataFrame(stats)


def main():
    parser = argparse.ArgumentParser(description="Data Profiling Tool")
    parser.add_argument("input_file", help="Đường dẫn đến file dữ liệu đầu vào (.csv, .xlsx, .xls)")
    parser.add_argument("output_file", help="Đường dẫn lưu file báo cáo đầu ra (.xlsx)")

    args = parser.parse_args()
    input_path = args.input_file
    output_path = args.output_file

    if not os.path.exists(input_path):
        print(f"[LỖI] Không tìm thấy file đầu vào: '{input_path}'")
        sys.exit(1)

    try:
        print(f"Đang đọc dữ liệu từ: {input_path}...")
        if input_path.lower().endswith('.csv'):
            df = pd.read_csv(input_path)
        elif input_path.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(input_path)
        else:
            print("[LỖI] Định dạng file không được hỗ trợ. Vui lòng sử dụng .csv hoặc .xlsx")
            sys.exit(1)
    except Exception as e:
        print(f"[LỖI] Quá trình đọc file gặp sự cố: {e}")
        sys.exit(1)

    print(f"Đã tải thành công tập dữ liệu: {len(df)} dòng, {len(df.columns)} cột.")

    profiler = DataProfiler(df)

    # regex_dict: chỉ cần khai báo khi có quy tắc nghiệp vụ cụ thể (vd mã đơn hàng phải là \d+).
    # Nếu không khai báo, Validity sẽ tự tính theo tỉ lệ parse số/ngày thành công (Structure Discovery).
    regex_dict = {}

    print("Đang tiến hành phân tích Profile dữ liệu...")
    df1 = profiler.get_basic_stats()
    df2 = profiler.get_numeric_stats(regex_dict)
    df3 = profiler.get_pattern_and_outliers()
    df4 = profiler.get_data_abstraction()

    try:
        print(f"Đang xuất báo cáo ra file: {output_path}...")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name='1. Thống kê cơ bản', index=False)
            df2.to_excel(writer, sheet_name='2. Validity & Numeric', index=False)
            df3.to_excel(writer, sheet_name='3. Patterns & Outliers', index=False)
            df4.to_excel(writer, sheet_name='4. Data Abstraction', index=False)

        wb = load_workbook(output_path)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(left=Side(style='thin', color='BDC3C7'),
                             right=Side(style='thin', color='BDC3C7'),
                             top=Side(style='thin', color='BDC3C7'),
                             bottom=Side(style='thin', color='BDC3C7'))

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    cell.border = thin_border

                    if sheet == '1. Thống kê cơ bản' and cell.column in [4, 6, 7] and cell.row > 1:
                        if cell.value is not None: cell.number_format = '0.00%'
                    if sheet == '2. Validity & Numeric' and cell.column == 3 and cell.row > 1:
                        if cell.value is not None: cell.number_format = '0.00%'

                ws.column_dimensions[column].width = min(max_length + 2, 40)
                ws[f'{column}1'].fill = header_fill
                ws[f'{column}1'].font = header_font
                ws[f'{column}1'].alignment = Alignment(horizontal='center', vertical='center')

        wb.save(output_path)
        print(f"[THÀNH CÔNG] Báo cáo đã được lưu tại: {output_path}")

    except Exception as e:
        print(f"[LỖI] Quá trình xuất file gặp sự cố: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()